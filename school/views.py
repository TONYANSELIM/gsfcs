from django.shortcuts import render, redirect
from django.http import HttpResponseForbidden
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.http import HttpResponseBadRequest
from django.http import JsonResponse
import qrcode
from io import BytesIO
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.conf import settings
from django.urls import reverse
import csv
import openpyxl
from django.contrib.auth import login, authenticate, logout
from .models import SchoolAdmin, Profile, SFacility, Maintanance, Announcement,DSchool,MAINTENANCE_CATEGORY
from .forms import UserForm, UserUpdateForm, ProfileUpdateForm, SchoolAdminForm, UserLoginForm, SFacilityForm, MaintananceForm, AnnouncementForm, CustomPasswordChangeForm,DSchoolForm,CommentForm
from django.db.models import Q
from django.contrib.auth.hashers import make_password
from .decorators import staff_required, active_user_required
from django.contrib import messages
import logging


logger = logging.getLogger(__name__)

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            if user.is_active:
                if user.is_staff:
                    login(request, user)
                    logger.info(f"Staff user {username} logged in successfully.")
                    messages.success(request, f"Welcome back, {username}.")
                    return redirect('school-index')  # Redirect to school index for staff users
                else:
                    login(request, user)
                    logger.info(f"Active user {username} logged in successfully.")
                    messages.success(request, f"Welcome back, {username}.")
                    return redirect('school-schooladmin_index')  # Redirect to admin index for active non-staff users
            else:
                logger.warning(f"Inactive account attempted login: {username}")
                messages.warning(request, 'Your account is inactive.')
                return render(request, 'districtdash/login.html')
        else:
            logger.warning(f"Failed login attempt with username: {username}")
            messages.error(request, 'Username or Password is incorrect')
            return render(request, 'districtdash/login.html')
    return render(request, 'districtdash/login.html')

def logout_view(request):
    logout(request)
    return redirect('district-logout')

@login_required
def view_comments(request):
    query = request.GET.get('q')
    school_id = request.GET.get('school')
    announcements = Announcement.objects.all().order_by('-TimePosted')
    
    if query:
        announcements = announcements.filter(Title__icontains=query)
    if school_id:
        announcements = announcements.filter(DSchoolID=school_id)
    
    schools = DSchool.objects.all()
    
    return render(request, 'schooladmin/addcomment.html', {
        'announcements': announcements,
        'schools': schools,
        'query': query,
        'selected_school': school_id
    })



@login_required
@staff_required
def download_payment_records(request):
    facility_q = request.GET.get('facility_q', '')
    school_id = request.GET.get('school', '')
    category = request.GET.get('category', '')

    maintenances = Maintanance.objects.filter(receipt__isnull=False).select_related('facility__school')

    if facility_q:
        maintenances = maintenances.filter(facility__FacilityName__icontains=facility_q)
    
    if school_id:
        maintenances = maintenances.filter(facility__school__DSchoolID=school_id)

    if category:
        maintenances = maintenances.filter(category=category)

    # Create an Excel workbook and add a worksheet.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment Records"

    # Add headers
    headers = ['#', 'Picture', 'Facility', 'Category', 'Facility Number', 'School', 'Status']
    ws.append(headers)

    # Add data
    for index, maintenance in enumerate(maintenances, start=1):
        row = [
            index,
            maintenance.MaintanancePicture.url,
            maintenance.facility.FacilityName,
            maintenance.get_category_display(),
            maintenance.facility.generate_facility_number(),
            maintenance.facility.school.DSchoolName,
            "Paid" if maintenance.confirmed else "Pending",
        ]
        ws.append(row)

    # Set the content type and filename
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=payment_records.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response

@login_required
def generate_pdf(request):
    facilities = SFacility.objects.all().select_related('school')

    template_path = 'schooldash/facilities_pdf.html'
    context = {'facilities': facilities}
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="facilities.pdf"'
    
    template = get_template(template_path)
    html = template.render(context)
    
    # Create PDF using pisa
    pisa_status = pisa.CreatePDF(html, dest=response)
    
    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')
    
    return response

@login_required
@staff_required
def download_maintenance_records(request):
    maintenances = Maintanance.objects.all()  # Adjust this queryset as per your filter needs

    # Create an Excel workbook and add a worksheet.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Maintenance Records"

    # Add headers
    headers = ['Facility Name', 'Category', 'Facility Number', 'School Name', 'Maintenance Description']
    ws.append(headers)

    # Add data
    for maintenance in maintenances:
        ws.append([
            maintenance.facility.FacilityName,
            maintenance.get_category_display(),
            maintenance.facility.generate_facility_number(),
            maintenance.facility.school.DSchoolName,
            maintenance.MaintananceDescription,
        ])

    # Set the content type and filename
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=maintenance_records.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response

@login_required
@active_user_required
def download_facilities_excel(request):
    facilities = SFacility.objects.all()

    # Create an Excel workbook and add a worksheet.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facilities List"

    # Add headers
    headers = ['Facility Name', 'School Name', 'Category', 'Facility Number']
    ws.append(headers)

    # Add data
    for facility in facilities:
        ws.append([
            facility.FacilityName,
            facility.school.DSchoolName,
            facility.get_FacilityCategory_display(),
            facility.generate_facility_number(),
        ])

    # Set the content type and filename
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=facilities_list.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response

@login_required
@active_user_required
def download_maintenance_excel(request):
    maintenances = Maintanance.objects.filter(confirmed=False)

    # Create an Excel workbook and add a worksheet.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Maintenance Records"

    # Add headers
    headers = ['Facility Name', 'Category', 'Amount', 'Facility Number', 'Description']
    ws.append(headers)

    # Add data
    for maintenance in maintenances:
        ws.append([
            maintenance.facility.FacilityName,
            maintenance.get_category_display(),
            maintenance.amount,
            maintenance.facility.generate_facility_number(),
            maintenance.MaintananceDescription,
        ])

    # Set the content type and filename
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=maintenance_records.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response

@login_required
@active_user_required
def download_paymentstatus_excel(request):
    facility_q = request.GET.get('facility_q', '')
    category = request.GET.get('category', '')

    confirmed_maintenances = Maintanance.objects.filter(confirmed=True).select_related('facility__school')

    if facility_q:
        confirmed_maintenances = confirmed_maintenances.filter(
            facility__FacilityName__icontains=facility_q
        )

    if category:
        confirmed_maintenances = confirmed_maintenances.filter(
            category=category
        )

    # Create an Excel workbook and add a worksheet.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Maintenance Payment Status"

    # Add headers
    headers = ['Facility Name', 'Category', 'Facility Number', 'School', 'Status', 'Description']
    ws.append(headers)

    # Add data
    for maintenance in confirmed_maintenances:
        ws.append([
            maintenance.facility.FacilityName,
            maintenance.get_category_display(),
            maintenance.facility.generate_facility_number(),
            maintenance.facility.school.DSchoolName,
            "Paid" if maintenance.receipt else "Pending",
            maintenance.MaintananceDescription,
        ])

    # Set the content type and filename
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=maintenance_payment_status.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response


@login_required
@staff_required
def index(request):
    items = SchoolAdmin.objects.all()
    items_count = items.count()
    schools_count = DSchool.objects.count()  # Count of schools
    facilities_count = SFacility.objects.count()  # Count of facilities
    maintanance_count = Maintanance.objects.filter(confirmed=False).count()  # Count of pending maintenances
    payment_count = Maintanance.objects.filter(receipt__isnull=False).count()  # Count of payments
    announcement_count = Announcement.objects.count()  # Count of announcements

    context = {
        'items': items,
        'items_count': items_count,
        'schools_count': schools_count,
        'facilities_count': facilities_count,  # Add facilities count to context
        'maintanance_count': maintanance_count,  # Add maintanance count to context
        'payment_count': payment_count,  # Add payment count to context
        'announcement_count': announcement_count,  # Add announcement count to context
    }
    logger.info(f"User {request.user.username} accessed the index page.")
    return render(request, 'schooldash/index.html', context)





@login_required
@staff_required
def addadmin(request):
    items = SchoolAdmin.objects.all()
    items_count = items.count()

    search_query = request.GET.get('search')
    if search_query:
        items = items.filter(
            Q(SchoolAdminID__icontains=search_query) |
            Q(AdminMname__icontains=search_query) |
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(AdminSname__icontains=search_query) |
            Q(AdminPhoneNumber__icontains=search_query)
        )

    user_form = UserForm()
    admin_form = SchoolAdminForm()
    
    if request.method == 'POST':
        user_form = UserForm(request.POST)
        admin_form = SchoolAdminForm(request.POST)
        if user_form.is_valid() and admin_form.is_valid():
            user_instance = user_form.save(commit=False)
            user_instance.password = make_password(str(request.POST['password']))
            user_instance.save()
            
            admin_instance = admin_form.save(commit=False)
            admin_instance.user = user_instance
            admin_instance.save()
            
            logger.info(f"Admin {admin_instance.AdminMname} created successfully.")
            return redirect('school-addadmin')
        else:
            logger.error("Errors in user or admin form.")
            print(user_form.errors)
            print(admin_form.errors)

    context = {
        'items': items,
        'user_form': user_form,
        'admin_form': admin_form,
        'items_count': items_count,
    }
    return render(request, 'schooldash/addadmin.html', context)


@login_required
def addschool(request):
    schools = DSchool.objects.all()
    schools_count = schools.count()

    if request.method == 'POST':
        form = DSchoolForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('school-addschool')
    else:
        form = DSchoolForm()

    q = request.GET.get('q')
    if q:
        schools = schools.filter(DSchoolName__icontains=q)

    context = {
        'form': form,
        'schools': schools,
        'search_query': q,
        'schools_count':schools_count,
    }
    return render(request, 'schooldash/addschool.html', context)

@login_required
def addschool_edit(request, pk):
    school = get_object_or_404(DSchool, pk=pk)
    if request.method == 'POST':
        form = DSchoolForm(request.POST, instance=school)
        if form.is_valid():
            form.save()
            return redirect('school-addschool')  # Redirect to the list page
    else:
        form = DSchoolForm(instance=school)

    context = {
        'form': form,
        'school': school,
    }
    return render(request, 'schooldash/addschool_edit.html', context)

@login_required
def addschool_delete(request, pk):
    school = get_object_or_404(DSchool, pk=pk)

    if request.method == 'POST':
        school.delete()
        return redirect('school-addschool')  # Redirect to the list page

    context = {
        'school': school,
    }
    return render(request, 'schooldash/addschool_delete.html', context)



@login_required
@active_user_required
def maintanance(request):
    if request.method == 'POST':
        form = MaintananceForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('schooladmin-maintanance')
    else:
        form = MaintananceForm()
    
    # Handle search query
    facility_q = request.GET.get('facility_q', '')
    if facility_q:
        maintenances = Maintanance.objects.filter(
            Q(facility__FacilityName__icontains=facility_q)
        )
    else:
        maintenances = Maintanance.objects.all()
    

    context = {
        'form': form,
        'maintenances': maintenances,
        'facility_q': facility_q,
    }
    return render(request, 'schooladmin/maintanance.html', context)


@login_required
@active_user_required
def edit_maintenance(request, maintenance_id):
    maintenance = get_object_or_404(Maintanance, pk=maintenance_id)
    if request.method == 'POST':
        form = MaintananceForm(request.POST, request.FILES, instance=maintenance)
        if form.is_valid():
            form.save()
            return redirect('schooladmin-maintanance')
    else:
        form = MaintananceForm(instance=maintenance)
    
    context = {
        'form': form,
        'maintenance': maintenance,
    }
    return render(request, 'schooladmin/edit_maintanance.html', context)

@login_required
@active_user_required
def delete_maintenance(request, maintenance_id):
    maintenance = get_object_or_404(Maintanance, pk=maintenance_id)
    if request.method == 'POST':
        maintenance.delete()
        return redirect('schooladmin-maintanance')
    
    context = {
        'maintenance': maintenance,
    }
    return render(request, 'schooladmin/delete_maintanance.html', context)

@login_required
@staff_required
def maintanance2(request):
    facility_q = request.GET.get('facility_q', '')
    school_id = request.GET.get('school', '')
    category = request.GET.get('category', '')
    form = MaintananceForm()

    maintenances = Maintanance.objects.filter(confirmed=False).select_related('facility__school')

    if facility_q:
        maintenances = maintenances.filter(facility__FacilityName__icontains=facility_q)
    
    if school_id:
        maintenances = maintenances.filter(facility__school__DSchoolID=school_id)

    if category:
        maintenances = maintenances.filter(category=category)

    confirmed_maintenances = Maintanance.objects.filter(confirmed=True).select_related('facility__school')

    schools = DSchool.objects.all()

    if request.method == 'POST':
        maintenance_id = request.POST.get('maintenance_id')
        maintenance = Maintanance.objects.get(pk=maintenance_id)
        maintenance.confirmed = True
        maintenance.save()
        return JsonResponse({'status': 'success', 'maintenance_id': maintenance_id})

    context = {
        'form': form,
        'maintenances': maintenances,
        'confirmed_maintenances': confirmed_maintenances,
        'facility_q': facility_q,
        'schools': schools,
        'school_id': school_id,
        'categories': MAINTENANCE_CATEGORY,
        'selected_category': category,
    }
    return render(request, 'schooldash/maintanance2.html', context)





@login_required
@staff_required
def payment(request):
    facility_q = request.GET.get('facility_q', '')
    school_id = request.GET.get('school', '')
    category = request.GET.get('category', '')
    form = MaintananceForm()

    maintenances = Maintanance.objects.filter(receipt__isnull=False).select_related('facility__school')

    if facility_q:
        maintenances = maintenances.filter(
            facility__FacilityName__icontains=facility_q
        )
    
    if school_id:
        maintenances = maintenances.filter(
            facility__school__DSchoolID=school_id
        )

    if category:
        maintenances = maintenances.filter(
            category=category
        )

    schools = DSchool.objects.all()

    context = {
        'form': form,
        'maintenances': maintenances,
        'facility_q': facility_q,
        'schools': schools,
        'school_id': school_id,
        'categories': MAINTENANCE_CATEGORY,
        'selected_category': category,
    }
    return render(request, 'schooldash/payment.html', context)




@login_required
@staff_required
def paymentreport(request):
    return render(request, 'schooldash/paymentreport.html')

@login_required
@staff_required
def viewmonthlyreport(request):
    return render(request, 'schooldash/viewmonthlyreport.html')

from django.shortcuts import render, redirect
from django.contrib.auth import login, authenticate
from .forms import UserForm

from django.shortcuts import render, redirect
from django.contrib.auth import login, authenticate
from .forms import UserForm
import logging

logger = logging.getLogger(__name__)

def register(request):
    if request.method == 'POST':
        logger.debug("POST request received")
        form = UserForm(request.POST)
        if form.is_valid():
            logger.debug("Form is valid")
            user = form.save()
            username = user.username
            password = form.cleaned_data['password']
            logger.info(f"User {username} created successfully.")
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                logger.info(f"User {username} authenticated successfully.")
                return redirect('school-schooladmin_index')
            else:
                logger.error("Authentication failed.")
                return redirect('district-login')
        else:
            logger.error("Form is not valid")
            messages.error(request, "Form is not valid. Please correct the errors.")
    else:
        logger.debug("GET request received")
        form = UserForm()
    
    context = {
        'form': form
    }
    return render(request, 'districtdash/register.html', context)

@login_required
@staff_required
def profile(request):
    try:
        profile = request.user.profile
    except Profile.DoesNotExist:
        profile = Profile.objects.create(staff=request.user)
    
    # Retrieve SchoolAdmin instance associated with the user
    try:
        school_admin = SchoolAdmin.objects.get(user=request.user)
        admin_phone_number = school_admin.AdminPhoneNumber
        school_name = school_admin.AdminSname
    except SchoolAdmin.DoesNotExist:
        admin_phone_number = None
        school_name = None
    
    context = {
        'profile': profile,
        'admin_phone_number': admin_phone_number,
        'school_name': school_name,
    }
    return render(request, 'districtdash/profile.html', context)

@login_required
@active_user_required
def profile2(request):
    try:
        profile = request.user.profile
    except Profile.DoesNotExist:
        profile = Profile.objects.create(staff=request.user)
    
    try:
        school_admin = SchoolAdmin.objects.get(user=request.user)
        admin_phone_number = school_admin.AdminPhoneNumber
        school_name = school_admin.AdminSname
    except SchoolAdmin.DoesNotExist:
        admin_phone_number = None
        school_name = None
    
    context = {
        'profile': profile,
        'admin_phone_number': admin_phone_number,
        'school_name': school_name,
    }
    return render(request, 'districtdash/profile2.html', context)

@login_required
@staff_required
def profile_update(request):
    try:
        profile = request.user.profile
    except Profile.DoesNotExist:
        profile = Profile.objects.create(staff=request.user)

    if request.method == 'POST':
        u_form = UserUpdateForm(request.POST, instance=request.user)
        p_form = ProfileUpdateForm(request.POST, request.FILES, instance=profile)
        if u_form.is_valid() and p_form.is_valid():
            u_form.save()
            p_form.save()
            return redirect('district-profile')
    else:
        u_form = UserUpdateForm(instance=request.user)
        p_form = ProfileUpdateForm(instance=profile)

    context = {
        'u_form': u_form,
        'p_form': p_form,
    }
    return render(request, 'districtdash/profile_update.html', context)

@login_required
@active_user_required
def profile_update2(request):
    try:
        profile = request.user.profile
    except Profile.DoesNotExist:
        profile = Profile.objects.create(staff=request.user)

    if request.method == 'POST':
        u_form = UserUpdateForm(request.POST, instance=request.user)
        p_form = ProfileUpdateForm(request.POST, request.FILES, instance=profile)
        if u_form.is_valid() and p_form.is_valid():
            u_form.save()
            p_form.save()
            return redirect('district-profile2')
    else:
        u_form = UserUpdateForm(instance=request.user)
        p_form = ProfileUpdateForm(instance=profile)

    context = {
        'u_form': u_form,
        'p_form': p_form,
    }
    return render(request, 'districtdash/profile_update2.html', context)

@login_required
@staff_required
def addadmin_delete(request, pk):
    item = SchoolAdmin.objects.get(id=pk)
    if request.method == 'POST':
        item.delete()
        return redirect('school-addadmin')
    context = {
        'item': item
    }
    return render(request, 'schooldash/addadmin_delete.html', context)

@login_required
@staff_required
def addadmin_edit(request, pk):
    item = SchoolAdmin.objects.get(id=pk)
    if request.method == 'POST':
        form = SchoolAdminForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return redirect('school-addadmin')
    else:
        form = SchoolAdminForm(instance=item)
    context = {
        'form': form,
    }
    return render(request, 'schooldash/addadmin_edit.html', context)


@login_required
def addannouncement(request):
    announcements = Announcement.objects.all().order_by('-TimePosted')
    announcement_count = announcements.count()
    if request.method == 'POST':
        form = AnnouncementForm(request.POST, request.FILES)
        if form.is_valid():
            announcement = form.save(commit=False)
            announcement.SchoolAdminID = request.user.schooladmin
            announcement.save()
            return redirect('school-addannouncement')
    else:
        form = AnnouncementForm()
    context = {
        'form': form,
        'announcement_count': announcement_count,
        'announcements': announcements,
    }
    return render(request, 'schooladmin/addannouncement.html', context)

def viewannouncement(request):
    query = request.GET.get('q')
    announcements = Announcement.objects.all().order_by('-TimePosted')

    if query:
        announcements = announcements.filter(Q(DSchoolID__DSchoolName__icontains=query))

    schools = DSchool.objects.all()

    return render(request, 'schooladmin/viewannouncement.html', {
        'announcements': announcements,
        'schools': schools,
        'query': query,
    })
    

@login_required
@active_user_required
def editannouncement(request, pk):
    announcement = get_object_or_404(Announcement, pk=pk)
    if request.method == 'POST':
        form = AnnouncementForm(request.POST, request.FILES, instance=announcement)
        if form.is_valid():
            form.save()
            return redirect('school-addannouncement')
    else:
        form = AnnouncementForm(instance=announcement)
    return render(request, 'schooladmin/editannouncement.html', {'form': form})

@login_required
@active_user_required
def deleteannouncement(request, pk):
    announcement = get_object_or_404(Announcement, pk=pk)
    if request.method == 'POST':
        announcement.delete()
        return redirect('school-addannouncement')
    return render(request, 'schooladmin/confirm_delete.html', {'object': announcement})


@login_required
@active_user_required
def addfacility(request):
    facilities = SFacility.objects.all()

    # Handle search functionality
    query = request.GET.get('q')
    if query:
        facilities = facilities.filter(Q(FacilityName__icontains=query) | Q(school__DSchoolName__icontains=query))

    if request.method == 'POST':
        form = SFacilityForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('school-addfacility')  
    else:
        form = SFacilityForm()

    schools = DSchool.objects.all()

    context = {
        'form': form,
        'facilities': facilities,
        'schools': schools,
        'query': query,  # Pass query for search input persistence
    }
    return render(request, 'schooladmin/addfacility.html', context)

@login_required
@active_user_required
def editfacility(request, pk):
    facility = get_object_or_404(SFacility, pk=pk)
    if request.method == 'POST':
        form = SFacilityForm(request.POST, request.FILES, instance=facility)
        if form.is_valid():
            form.save()
            return redirect('school-addfacility')
    else:
        form = SFacilityForm(instance=facility)
    
    schools = DSchool.objects.all()  # Query all schools
    
    context = {
        'form': form,
        'facility': facility,
        'schools': schools,  # Pass schools queryset to context
    }
    return render(request, 'schooladmin/editfacility.html', context)


@login_required
@active_user_required
def deletefacility(request, pk):
    facility = get_object_or_404(SFacility, pk=pk)
    if request.method == 'POST':
        facility.delete()
        return redirect('school-addfacility')
    
    context = {
        'facility': facility,
    }
    return render(request, 'schooladmin/deletefacility.html', context)




def generate_qr_code(request, facility_id):
    facility = get_object_or_404(SFacility, pk=facility_id)
    facility_number = facility.generate_facility_number()
    
    # Generate QR code
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(facility_number)
    qr.make(fit=True)

    img = qr.make_image(fill='black', back_color='white')
    
    # Save the QR code to a BytesIO object
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    buffer.seek(0)
    
    # Return the image as a response
    return HttpResponse(buffer, content_type='image/png')


@login_required
@staff_required
def viewitems(request):
    query = request.GET.get('q')
    school_id = request.GET.get('school')

    facilities = SFacility.objects.all()

    # Apply search query if provided
    if query:
        facilities = facilities.filter(Q(FacilityName__icontains=query))

    # Apply school filter if provided
    if school_id:
        facilities = facilities.filter(school_id=school_id)

    # Fetch all schools for the dropdown filter
    schools = DSchool.objects.all()

    context = {
        'facilities': facilities,
        'query': query,
        'school_id': school_id,
        'schools': schools,  # Pass schools queryset to template
    }
    return render(request, 'schooldash/viewitems.html', context)

@login_required
@active_user_required
def addmaintainance(request):
    return render(request, 'schooladmin/addmaintainance.html')


@login_required
@active_user_required
def admin_paymentstatus(request):
    facility_q = request.GET.get('facility_q', '')
    category = request.GET.get('category', '')
    form = MaintananceForm()

    confirmed_maintenances = Maintanance.objects.filter(confirmed=True).select_related('facility__school')

    if facility_q:
        confirmed_maintenances = confirmed_maintenances.filter(
            facility__FacilityName__icontains=facility_q
        )

    if category:
        confirmed_maintenances = confirmed_maintenances.filter(
            category=category
        )

    schools = DSchool.objects.all()

    if request.method == 'POST':
        maintenance_id = request.POST.get('maintenance_id')
        maintenance = get_object_or_404(Maintanance, pk=maintenance_id)

        if 'receipt' in request.FILES:
            maintenance.receipt = request.FILES['receipt']
            maintenance.save()
        else:
            maintenance.confirmed = True
            maintenance.save()

        return redirect('school-admin_paymentstatus')

    context = {
        'form': form,
        'confirmed_maintenances': confirmed_maintenances,
        'facility_q': facility_q,
        'schools': schools,
        'categories': MAINTENANCE_CATEGORY,
        'selected_category': category,
    }
    return render(request, 'schooladmin/admin_paymentstatus.html', context)



@login_required
@active_user_required
def schooladmin_index(request):
    facilities_count = SFacility.objects.count()
    maintanance_count = Maintanance.objects.count()
    announcement_count = Announcement.objects.count()
    payment_count = Maintanance.objects.filter(confirmed=True).count()  # Count confirmed maintenances
    return render(request, 'schooladmin/schooladmin_index.html', {
        'facilities_count': facilities_count,
        'maintanance_count': maintanance_count,
        'announcement_count': announcement_count,
        'payment_count': payment_count,
    })




@login_required
@staff_required
def change_password_view(request):
    if request.method == 'POST':
        form = CustomPasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Your password was successfully updated!')
            return redirect('change_password')  # Redirect to change password page
    else:
        form = CustomPasswordChangeForm(request.user)
    
    return render(request, 'districtdash/change_password.html', {'form': form})

@login_required
@active_user_required
def change_password_view2(request):
    if request.method == 'POST':
        form = CustomPasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Your password was successfully updated!')
            return redirect('change_password2')  # Redirect to change password page
    else:
        form = CustomPasswordChangeForm(request.user)
    
    return render(request, 'districtdash/change_password2.html', {'form': form})